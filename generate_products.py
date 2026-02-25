import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.makedirs('productos', exist_ok=True)

# Styles
header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
input_fill = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid') # Yellow
result_fill = PatternFill(start_color='BBF7D0', end_color='BBF7D0', fill_type='solid') # Green
bold_font = Font(bold=True)
align_center = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_cell(sheet, row, col, value, fill=None, font=None, align=None, border=None, number_format=None):
    cell = sheet.cell(row=row, column=col)
    cell.value = value
    if fill: cell.fill = fill
    if font: cell.font = font
    if align: cell.alignment = align
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell


def create_e030_template():
    wb = Workbook()
    
    # ------------------ HOJA DATOS ------------------
    ws_datos = wb.active
    ws_datos.title = "1. Datos"
    ws_datos.column_dimensions['B'].width = 35
    ws_datos.column_dimensions['C'].width = 20
    ws_datos.column_dimensions['D'].width = 40
    
    set_cell(ws_datos, 2, 2, "PARÁMETROS DEL PROYECTO (NORMA E.030)", font=Font(bold=True, size=14))
    
    data_inputs = [
        ("Zona Sísmica (1 a 4)", 4, "input"),
        ("Tipo de Suelo (S0 a S3)", "S2", "input"),
        ("Categoría Edificación (A a D)", "B", "input"),
        ("Sistema Estructural", "Pórticos C.A.", "input"),
        ("Número de Pisos", 5, "input"),
        ("Altura Promedio Piso (m)", 3.0, "input"),
    ]
    
    for i, (label, val, ctype) in enumerate(data_inputs):
        r = 4 + i
        set_cell(ws_datos, r, 2, label, font=bold_font, border=thin_border)
        set_cell(ws_datos, r, 3, val, fill=input_fill, align=align_center, border=thin_border)
        
    set_cell(ws_datos, 12, 2, "PESO POR PISO (Ton)", font=Font(bold=True, size=12))
    set_cell(ws_datos, 13, 2, "Piso", fill=header_fill, font=header_font, align=align_center, border=thin_border)
    set_cell(ws_datos, 13, 3, "Peso (P)", fill=header_fill, font=header_font, align=align_center, border=thin_border)
    
    for p in range(1, 6):
        r = 13 + p
        set_cell(ws_datos, r, 2, f"Piso {p}", align=align_center, border=thin_border)
        set_cell(ws_datos, r, 3, 500, fill=input_fill, align=align_center, border=thin_border)

    set_cell(ws_datos, 20, 2, "Peso Total de la Edificación =", font=bold_font, border=thin_border)
    set_cell(ws_datos, 20, 3, "=SUM(C14:C18)", fill=result_fill, font=bold_font, align=align_center, border=thin_border)
    
    # Tabla de referencia
    set_cell(ws_datos, 4, 5, "REFERENCIA NORMA E.030", font=Font(bold=True))
    ref_data = [
        ("Factores Z:", "Z1: 0.10 | Z2: 0.25 | Z3: 0.35 | Z4: 0.45"),
        ("Categoría U:", "A=1.5 | B=1.3 | C=1.0"),
        ("Suelo S:", "Depende de Zona y Perfil. Ver Tabla N°3 de la norma."),
        ("Sistemas R:", "Pórticos=8 | Dual=7 | Muros=6 | Albañilería=3")
    ]
    for i, (lbl, desc) in enumerate(ref_data):
        set_cell(ws_datos, 5+i, 5, lbl, font=bold_font)
        set_cell(ws_datos, 5+i, 6, desc)

    
    # ------------------ HOJA CÁLCULO ------------------
    ws_calc = wb.create_sheet("2. Cálculo V Basal")
    ws_calc.column_dimensions['B'].width = 25
    ws_calc.column_dimensions['C'].width = 15
    ws_calc.column_dimensions['D'].width = 50
    
    set_cell(ws_calc, 2, 2, "CÁLCULO DEL CORTANTE BASAL V = (ZUCS/R)*P", font=Font(bold=True, size=14))
    
    calc_params = [
        ("Z (Factor de Zona)", 0.45, "Valor asignado de tabla Z"),
        ("U (Factor de Uso)", 1.30, "Categoría del edificio"),
        ("S (Factor de Suelo)", 1.05, "De tabla de suelo y zona"),
        ("CT (Periodo Fund.)", 35, "35 (pórticos), 45 (muros), 60 (acero)"),
        ("R (Coef. Reducción)", 8, "De tabla de sistema estructural"),
    ]
    
    for i, (label, val, comment) in enumerate(calc_params):
        r = 4 + i
        set_cell(ws_calc, r, 2, label, font=bold_font, border=thin_border)
        set_cell(ws_calc, r, 3, val, fill=input_fill, align=align_center, border=thin_border)
        set_cell(ws_calc, r, 4, comment)
        
    set_cell(ws_calc, 11, 2, "Tp (Periodo Suelo)", font=bold_font, border=thin_border)
    set_cell(ws_calc, 11, 3, 0.6, fill=input_fill, align=align_center, border=thin_border)
    
    set_cell(ws_calc, 13, 2, "L (Longitud Tot) m", font=bold_font, border=thin_border)
    set_cell(ws_calc, 13, 3, "='1. Datos'!C8*'1. Datos'!C9", fill=result_fill, align=align_center, border=thin_border, number_format='0.00')
    
    set_cell(ws_calc, 14, 2, "T = hn/CT", font=bold_font, border=thin_border)
    set_cell(ws_calc, 14, 3, "=C13/C7", fill=result_fill, align=align_center, border=thin_border, number_format='0.00')
    
    set_cell(ws_calc, 15, 2, "C (Factor Amp.)", font=bold_font, border=thin_border)
    set_cell(ws_calc, 15, 3, '=IF(C14<C11, 2.5, 2.5*(C11/C14))', fill=result_fill, align=align_center, border=thin_border, number_format='0.00')
    set_cell(ws_calc, 15, 4, "Calculado dinámicamente según T")
    
    set_cell(ws_calc, 17, 2, "Peso Total P (Ton)", font=bold_font, border=thin_border)
    set_cell(ws_calc, 17, 3, "='1. Datos'!C20", fill=result_fill, align=align_center, border=thin_border, number_format='0.00')

    set_cell(ws_calc, 19, 2, "V BASAL (Ton)", font=Font(bold=True, size=12, color='FFFFFF'), fill=PatternFill(start_color='C53030', end_color='C53030', fill_type='solid'), align=align_center, border=thin_border)
    set_cell(ws_calc, 19, 3, "=(C4*C5*C15*C6/C8)*C17", font=Font(bold=True, size=12), fill=result_fill, align=align_center, border=thin_border, number_format='0.00')
    
    # ------------------ HOJA DISTRIBUCIÓN ------------------
    ws_res = wb.create_sheet("3. Distribución Fx")
    set_cell(ws_res, 2, 2, "DISTRIBUCIÓN DE CORTANTE BASAL POR PISO", font=Font(bold=True, size=14))
    
    headers = ["Piso", "Peso (P) Ton", "Altura (h) m", "P*h", "Factor (% V)", "Fuerza Fx (Ton)", "Cortante Vi (Ton)"]
    for i, h in enumerate(headers):
        set_cell(ws_res, 4, 2+i, h, fill=header_fill, font=header_font, align=align_center, border=thin_border)
        ws_res.column_dimensions[get_column_letter(2+i)].width = 16
        
    ws_res.protection.sheet = True
    wb.save('productos/plantilla-diseno-sismico-e030.xlsx')


def create_metrados_template():
    wb = Workbook()
    
    # ------------------ CONCRETO ------------------
    ws = wb.active
    ws.title = "1. Concreto"
    
    for k, w in [('B',10),('C',35),('D',10),('E',10),('F',10),('G',10),('H',10),('I',15)]:
        ws.column_dimensions[k].width = w
        
    set_cell(ws, 2, 2, "METRADO DE CONCRETO (OE.02.03 CONCRETO ARMADO)", font=Font(bold=True, size=14))
    
    headers = ["Ítem", "Descripción", "N° Veces", "Largo (m)", "Ancho (m)", "Alto (m)", "Parcial (m3)", "Total (m3)"]
    for i, h in enumerate(headers):
        set_cell(ws, 4, 2+i, h, fill=header_fill, font=header_font, align=align_center, border=thin_border)
        
    items = [
        ("01.01", "Zapatas Z-1", 4, 1.5, 1.5, 0.6),
        ("01.02", "Columnas C-1", 8, 0.3, 0.3, 3.2),
        ("01.03", "Vigas VP-1", 4, 5.0, 0.25, 0.5),
        ("01.04", "Losa Aligerada h=0.2m", 1, 10.0, 8.0, 0.2)
    ]
    
    for i, rowdata in enumerate(items):
        r = 5 + i
        set_cell(ws, r, 2, rowdata[0], align=align_center, border=thin_border)
        set_cell(ws, r, 3, rowdata[1], border=thin_border)
        for j in range(4):
            set_cell(ws, r, 4+j, rowdata[2+j], fill=input_fill, align=align_center, border=thin_border)
        
        # Parcial
        set_cell(ws, r, 8, f"=D{r}*E{r}*F{r}*G{r}", align=align_center, border=thin_border, number_format='0.00')
        # Total
        set_cell(ws, r, 9, f"=H{r}" if i==0 else f"=I{r-1}+H{r}", fill=result_fill, align=align_center, border=thin_border, number_format='0.00')

    set_cell(ws, 10, 8, "TOTAL VOLUMEN =", font=bold_font, align=align_center)
    set_cell(ws, 10, 9, "=SUM(H5:H8)", fill=result_fill, font=bold_font, align=align_center, border=thin_border, number_format='0.00 m3')

    # ------------------ ACERO ------------------
    ws2 = wb.create_sheet("2. Acero")
    set_cell(ws2, 2, 2, "METRADO DE ACERO ESTRUCTURAL (Kg)", font=Font(bold=True, size=14))
    
    headers2 = ["Elemento", "Ø (pulg)", "Factor (Kg/m)", "L unit (m)", "N° Piezas", "Long Total (m)", "Peso (Kg)"]
    for i, h in enumerate(headers2):
        set_cell(ws2, 4, 2+i, h, fill=PatternFill(start_color='B91C1C', end_color='B91C1C', fill_type='solid'), font=header_font, align=align_center, border=thin_border)
        ws2.column_dimensions[get_column_letter(2+i)].width = 15
        
    set_cell(ws2, 5, 2, "Columna C1 - Lng", border=thin_border)
    set_cell(ws2, 5, 3, "5/8", fill=input_fill, align=align_center, border=thin_border)
    set_cell(ws2, 5, 4, 1.552, align=align_center, border=thin_border) # factor
    set_cell(ws2, 5, 5, 4.10, fill=input_fill, align=align_center, border=thin_border)
    set_cell(ws2, 5, 6, 6, fill=input_fill, align=align_center, border=thin_border)
    
    set_cell(ws2, 5, 7, "=E5*F5", align=align_center, border=thin_border, number_format='0.00')
    set_cell(ws2, 5, 8, "=D5*G5", fill=result_fill, font=bold_font, align=align_center, border=thin_border, number_format='0.00')

    # ------------------ ENCOFRADO ------------------
    ws3 = wb.create_sheet("3. Encofrado")
    set_cell(ws3, 2, 2, "METRADO DE ENCOFRADO (m2)", font=Font(bold=True, size=14))

    # ------------------ RESUMEN ------------------
    ws4 = wb.create_sheet("4. Cuadro Resumen")
    set_cell(ws4, 2, 2, "RESUMEN DE METRADOS (PRESUPUESTO BASE)", font=Font(bold=True, size=14))
    set_cell(ws4, 4, 2, "Partida", fill=header_fill, font=header_font, border=thin_border)
    set_cell(ws4, 4, 3, "Unidad", fill=header_fill, font=header_font, border=thin_border)
    set_cell(ws4, 4, 4, "Total", fill=header_fill, font=header_font, border=thin_border)
    
    res = [("Concreto Armado", "m3", "='1. Concreto'!I10"),
           ("Encofrado Columnas", "m2", 120.5),
           ("Acero Estructural", "Kg", "='2. Acero'!H5")]
           
    for i, (p, u, t) in enumerate(res):
        r = 5 + i
        set_cell(ws4, r, 2, p, border=thin_border)
        set_cell(ws4, r, 3, u, align=align_center, border=thin_border)
        set_cell(ws4, r, 4, t, fill=result_fill, font=bold_font, align=align_center, border=thin_border, number_format='0.00')
        
    for k, w in [('B',30),('C',15),('D',20)]: ws4.column_dimensions[k].width = w

    wb.save('productos/plantilla-metrados-obra.xlsx')

if __name__ == '__main__':
    create_e030_template()
    create_metrados_template()
